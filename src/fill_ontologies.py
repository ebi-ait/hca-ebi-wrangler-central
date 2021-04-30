#!/usr/bin/env python3

"""
Requirements

python>=3.5
hca-ingest==0.6.11
xlrd>=1.0.0
pandas
openpyxl

Instructions of use and main algorithm

1. Algorithm

    a. Load the workbook with openpyxl tools. Name will be the arg passed as argv[1]
    b. Try to load pickled schemas, if not present at pickled_schemas.pkl,
        Load a SchemaTemplate object and save schemas to picked_schemas.pkl (Takes a while, but necessary)
    c. Parse each cell of the workbook below row 5
    d. When an ontology property is found (Row 4.endswith("text"), load the schema from the full qualified key to apply
    DB restrictions
    e. Query the OLS/HCAO with database restrictions
    f. If array delimiter `||` is detected in the cell, each term is automatically searched and the returned curations
      are concatenated in the spreadsheet.
    g. If ZOOMA option enabled, query the ZOOMA API for any available curation
    h. If `--keep` option specified, no existing ontologies will be overwritten.
    i. Return ontologies that match the search and ask the user which one is correct.
    j. Save spreadsheet with the same name + _ontologies

1. Instruction of use

    **Install requirements** first (See above)

    ```
    cd src/
    python3 fill_ontologies.py -s <path_to_spreadsheet> [OPTIONS -z -k]
    ```
    The script takes some time to get started (It's loading a SchemaTemplate object). After approx 30 secs/1 min, it
    will begin the search, with the following possible outputs:
    - `Found exact match for term 'x' (Ontology: 'y')`: The program has found an ontology that matches perfectly the
    spreadsheet value
    - `Found high confidence, HCA match for term 'x' (Ontology: y)`
    - `x matches found for your search term 'y' for field <spreadsheet text value>.`
        - Numbered list of matches in format (ZOOMA info show with option -z):
            - n. ontology_label - ontology_id. ZOOMA: confidence: (HIGH/GOOD/MEDIUM/LOW), source: <SOURCE>.
            - Enter the appropriate number or 'm' for manual input or 'none' to skip this term
        - Input n: The tool will fill the spreadsheet with that ontology.
        - Input 'm': You'll be prompted to introduce the term that you want to look for
        - Input 'none': The ontology cells will be filled with empty strings
    - No ontology was found for this term. Please input it manually: Input the term that you want to search.
"""

import requests as rq
import argparse
import openpyxl
from ingest.template.schema_template import SchemaTemplate
from openpyxl.utils import get_column_letter
import pickle
import re
import os
import sys


def define_parser():
    parser = argparse.ArgumentParser(description="Parser for the arguments")
    parser.add_argument("--spreadsheet", "-s", action="store", dest="wb_path", type=str, help="Input spreadsheet name")
    parser.add_argument("--zooma", "-z", action="store_true", dest="zooma",
                        help="If specified, will enable curation options by querying ZOOMA & OLS. Default OLS only.")
    parser.add_argument("--keep", "-k", action="store_true", dest="keep",
                        help="If specified, will keep any existing ontology curations in the sheet. Default will overwrite.")
    return parser


def get_iri(classes, iri={}):
    # Get an iri from given classes
    for ontology_class in classes:
        ontology_name = ontology_class.split(":")[0]
        if ontology_name in iri:
            continue
        else:
            request = rq.get("https://www.ebi.ac.uk/ols/api/terms?id={}".format(ontology_class))
            response = request.json()
            iri[ontology_name] = "/".join(response["_embedded"]["terms"][0]["iri"].split("/")[:-1])
    return iri


def get_ontology_schemas(key, json_schemas):
    # Get ontology graph restrictions from a given key
    key = key.split(".")
    try:
        for schema in json_schemas:
            if "name" not in schema:
                continue
            elif key[0] == schema["name"]:
                key.pop(0)
                while len(key) > 1:
                    if schema["properties"][key[0]]["type"] == "array":
                        if isinstance(schema["properties"][key[0]]["items"], list):
                            response = rq.get(schema["properties"][key[0]]["items"][0]["$ref"])
                        else:
                            response = rq.get(schema["properties"][key[0]]["items"]["$ref"])
                    else:
                        response = rq.get(schema["properties"][key[0]]["$ref"])
                    schema = response.json()
                    key.pop(0)
                return schema
    except KeyError as e:
        print(e)
        print("Could not find key: {} in schemas. Try updating your pickle by deleting pickled_schemas.pkl".format(key))
        sys.exit(0)


def get_schema_info(key, json_schemas):
    # Get schema info from a given key
    search_schema = get_ontology_schemas(key, json_schemas)
    search_ontologies = search_schema["properties"]["ontology"]["graph_restriction"]["ontologies"]
    classes = search_schema["properties"]["ontology"]["graph_restriction"]["classes"]
    ontologies = ",".join([ontology.split(":")[1] for ontology in search_ontologies]).lower()
    include_self = search_schema["properties"]["ontology"]["graph_restriction"]["include_self"]
    schema_info = {"classes": classes,
                   "ontologies": ontologies,
                   "include_self": include_self}
    return schema_info


def search_child_term(term, schema_info, iri={}):
    # Search OLS for ontologies based on string matching in ontologies determined by schema graph restriction
    ontology_response = []
    request_query = "https://ontology.archive.data.humancellatlas.org/api/search?q=" if "hcao" in schema_info['ontologies'] else \
                    "http://www.ebi.ac.uk/ols/api/search?q="
    if schema_info['include_self']:
        for ontology_class in schema_info['classes']:
            request = rq.get("http://www.ebi.ac.uk/ols/api/terms?id={}".format(ontology_class))
            response = request.json()
            if response["_embedded"]["terms"][0]["label"] == term:
                return {response["_embedded"]["terms"][0]["obo_id"]: response["_embedded"]["terms"][0]}, iri

    iri = get_iri(schema_info['classes'], iri)
    iri_query = ",".join([iri[ontology_class.split(":")[0]] + "/" + ontology_class.replace(":", "_") for ontology_class in schema_info['classes']])
    request = request_query + "{}&ontology={}&allChildrenOf={}".format(term, schema_info['ontologies'], iri_query)
    response = rq.get(request).json()
    if response["response"]["numFound"] != 0:
        ontology_response.extend(response["response"]["docs"])
    else:
        if "none" in term.lower():
            return [{"obo_id": "", "label": ""}], iri
        return None, iri
    ontology_dict = {ontology['obo_id']: ontology for ontology in ontology_response}
    return ontology_dict, iri


def search_zooma(term, schema_info):
    # Search ZOOMA datasource for curated ontologies
    # TODO: Update to production instance of zooma
    query = "http://snarf.ebi.ac.uk:8580/spot/zooma/v2/api/services/annotate?filter=preferred:[hca]&ontologies:[{}]&propertyValue={}".format(schema_info['ontologies'], term)
    zooma_response = rq.get(query)
    if zooma_response:
        response_json = zooma_response.json()
    else:
        print("No Zooma curations found, continuing to OLS search.")
        return None

    annotation_dict = {}
    for annotation in response_json:
        if annotation['provenance']['evidence'] == 'ZOOMA_INFERRED_FROM_CURATED':
            ols_response = rq.get('http://www.ebi.ac.uk/ols/api/terms/findByIdAndIsDefiningOntology?iri={}'.format(annotation['semanticTags'][0]))
            ols_json = ols_response.json()
            if "_embedded" not in ols_json.keys():
                ols_response = rq.get('http://www.ebi.ac.uk/ols/api/terms?iri={}'.format(annotation['semanticTags'][0]))
                ols_json = ols_response.json()
            try:
                zooma_dict = ols_json['_embedded']['terms'][0]
                zooma_dict['confidence'] = annotation['confidence']
                zooma_dict['source'] = annotation['derivedFrom']['provenance']['source']['name']
                annotation_dict[zooma_dict['obo_id']] = zooma_dict
            except KeyError as e:
                if e == "_embedded":
                    print("Failed to retrieve ontology information.")
                    return None
    return annotation_dict


def select_term(ontologies_dict, term, key, schema_info, zooma, known_iri={}, multi_flag=False):
    first_key = next(iter(ontologies_dict))
    # If there is a high confidence match from HCA, use it
    if "confidence" in ontologies_dict[first_key].keys() and ontologies_dict[first_key]["confidence"] == "HIGH" and \
            ontologies_dict[first_key]["source"] == "HCA":
        print("Found high confidence, HCA match for term {} (Ontology: {})".format(term, ontologies_dict[first_key]["obo_id"]))
        return ontologies_dict[first_key], known_iri
    # If there is an exact string match, use it
    elif ontologies_dict and ontologies_dict[first_key]["label"].lower() == term.lower() and not zooma:
        print("Found exact match for term {} (Ontology: {})".format(term, ontologies_dict[first_key]["obo_id"]))
        return ontologies_dict[first_key], known_iri
    else:
        if not ontologies_dict or len(ontologies_dict) == 0:
            term = input("No ontologies were found for the term (Cell value = {}, Key = {}). Please input it manually: "
                         .format(term, key))
            ontologies_dict, known_iri = search_child_term(term, schema_info)
            return select_term(ontologies_dict, term, key, schema_info, zooma, known_iri)
        # If multiple terms present, search each term
        if re.search("\|\|", term) and not multi_flag:
            print("Multiple terms detected for term: {} in field {}.".format(term, key))
            multi_ontology = []
            terms_to_search = term.split("||")
            for each_term in terms_to_search:
                each_ontologies_dict, known_iri = search_child_term(each_term, schema_info)
                if zooma:
                    zooma_ann_dict = search_zooma(each_term, schema_info)
                    if each_ontologies_dict and zooma_ann_dict:
                        each_ontologies_dict = {**each_ontologies_dict, **zooma_ann_dict}
                    elif not each_ontologies_dict and zooma_ann_dict:
                        each_ontologies_dict = zooma_ann_dict
                    elif not each_ontologies_dict and not zooma_ann_dict:
                        manual_term = input(
                            "Term '{}' was not found (Property = {}).\nPlease input it manually:".format(each_term, key))
                        each_ontologies_dict, known_iri = search_child_term(manual_term, schema_info)
                # TODO: extend only ontologies, not known iri
                multi_ontology.extend(select_term(each_ontologies_dict, each_term, key, schema_info, zooma, known_iri,
                                                  multi_flag=True))
            return multi_ontology, known_iri
        print("{} matches found for your search term '{}' for field {}.".format(len(ontologies_dict), term, key))
        i = 1
        for ontology, info in ontologies_dict.items():
            if "obo_id" not in info.keys():
                continue
            if "confidence" in info.keys():
                print("{}. {} - {}. ZOOMA: confidence: {}, source: {}.".format(i, info["label"], info["obo_id"],
                                                                               info["confidence"], info["source"]))
            else:
                print("{}. {} - {}".format(i, info["label"], info["obo_id"]))
            i += 1
        answer = input("Enter the appropriate number, 'm' for manual input, 'none' to skip this term, 'q' to save progess and stop program.\n")
        if answer.lower() == 'm':
            term = input("Please input the term manually: ")
            ontologies_dict, known_iri = search_child_term(term, schema_info)
            return select_term(ontologies_dict, term, key, schema_info, zooma, known_iri)
        elif answer.lower() == 'none' or answer.lower() == 'skip':
            blank_annotation = {"obo_id": "",
                                "label": ""}
            return blank_annotation, known_iri
        elif answer.lower() == 'exit' or answer.lower() == "q":
            raise KeyboardInterrupt
        else:
            dict_index_key = list(ontologies_dict.keys())[int(answer)-1]
            return ontologies_dict[dict_index_key], known_iri


def save_workbook(path, workbook, suffix="_ontologies"):
    # Save workbook with adjusted name to same location as input path.
    split_path = os.path.split(path)
    wb_name = split_path[-1]
    wb_dir = split_path[0]
    output_name = "".join(wb_name.split(".")[:-1]) + suffix + ".xlsx"
    workbook.save(os.path.join(wb_dir, output_name))
    print("Saved output to {}.".format(os.path.join(wb_dir, output_name)))


def parse_wb(file_path, wb, schema, zooma, keep):
    # Parses through workbook with nested for loops to iterate each tab, each row, then each column
    # When `.text` field found, search for ontologies, prompt user to select correct ontology.
    # Save workbook when finished or interrupted.
    known_terms = []
    known_ontologies = {}
    try:
        for sheet in wb:
            for row in sheet.rows:
                if list(row)[0].row < 5:
                    continue
                for cell in row:
                    if sheet["{}4".format(get_column_letter(cell.column))].value.endswith(".text") and cell.value \
                            and keep:  # If there is already an ontology curation, and keep enabled, skip curation
                        if sheet["{}{}".format(get_column_letter(cell.column + 1), cell.row)].value:
                            continue
                    if sheet["{}4".format(get_column_letter(cell.column))].value and \
                            sheet["{}4".format(get_column_letter(cell.column))].value.endswith(".text") and \
                            cell.value:  # If there is a filled `.text` cell attempt curation:
                        if cell.value in known_terms:  # If we already curated the term, use existing
                            sheet["{}{}".format(get_column_letter(cell.column + 1), cell.row)].value = known_ontologies[cell.value]["obo_id"]
                            sheet["{}{}".format(get_column_letter(cell.column + 2), cell.row)].value = known_ontologies[cell.value]["label"]
                            continue
                        programmatic_key = sheet["{}4".format(get_column_letter(cell.column))].value
                        schema_info = get_schema_info(programmatic_key, schema)
                        ontologies_dict, known_iri = search_child_term(cell.value, schema_info)  # Search OLS
                        if zooma:
                            zooma_ann_dict = search_zooma(cell.value, schema_info)  # Search ZOOMA
                            if ontologies_dict and zooma_ann_dict:  # Combine OLS and ZOOMA ontologies
                                ontologies_dict = {**ontologies_dict, **zooma_ann_dict}
                            elif not ontologies_dict and zooma_ann_dict:
                                ontologies_dict = zooma_ann_dict
                            elif not ontologies_dict and not zooma_ann_dict:
                                term = input("Term '{}' was not found (Property = {}).\nPlease input it manually:".format(cell.value, programmatic_key))
                                ontologies_dict, known_iri = search_child_term(term, schema_info)
                        if ontologies_dict:  # Select term from found ontologies
                            ontology, known_iri = select_term(ontologies_dict, cell.value, programmatic_key,
                                                              schema_info, zooma, known_iri)
                            known_terms.append(cell.value)
                            if isinstance(ontology, list):  # If multi ontology returned, concatenate values
                                ontologies = {}
                                ontologies["obo_id"] = "||".join([ontology_element["obo_id"] for ontology_element in ontology if "obo_id" in ontology_element])
                                ontologies["label"] = "||".join([ontology_label["label"] for ontology_label in ontology if "obo_id" in ontology_label])
                                ontology = ontologies
                            known_ontologies[cell.value] = ontology
                            sheet["{}{}".format(get_column_letter(cell.column + 1), cell.row)].value = known_ontologies[cell.value]["obo_id"]
                            sheet["{}{}".format(get_column_letter(cell.column + 2), cell.row)].value = known_ontologies[cell.value]["label"]
        save_workbook(file_path, wb)

    except KeyboardInterrupt:
        save_workbook(file_path, wb, "_partial_ontologies")
    sys.exit(0)


def main(args):
    wb = openpyxl.load_workbook(args.wb_path)
    try:
        with open('pickled_schemas.pkl', 'rb') as pickled_schemas:
            json_schemas = pickle.load(pickled_schemas)
    except FileNotFoundError:
        schema = SchemaTemplate(ingest_api_url="http://api.ingest.staging.archive.data.humancellatlas.org")
        with open('pickled_schemas.pkl', 'wb') as output:
            pickle.dump(schema.json_schemas, output)
        json_schemas = schema.json_schemas
    parse_wb(args.wb_path, wb, json_schemas, args.zooma, args.keep)


if __name__ == "__main__":
    parser = define_parser()
    args = parser.parse_args()
    main(args)
