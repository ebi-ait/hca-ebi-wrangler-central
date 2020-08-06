import argparse
import os
import requests as rq
from io import BytesIO
from typing import Tuple, List, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter

accession_mapping = {'sample': {'tab': ['Donor organism', 'Specimen from organism', 'Cell suspension'],
                     'fqk': '{}.biomaterial_core.biosamples_accession'},
                     'project': {'tab': ['Project'],
                                 'fqk': '{}.biostudies_accessions'},
                     'sequencingExperiment': {'tab': ['Sequence File'],
                                              'fqk': 'process.insdc_experiment.insdc_experiment_accession'},
                     'sequencingRun': {'tab': ['Sequence File'],
                                       'fqk': '{}.insdc_run_accessions'},
                     'study': {'tab': ['Project'],
                               'fqk': '{}.insdc_project_accessions'}
                     }


def search_fqk_in_sheet(sheet: Worksheet, fqk: str, row_number: int = 4):
    """
    Search for a Full Qualified Key (fqk) in the spreadsheet at a certain row and return the column letter.

    :param sheet: Worksheet to search in
    :param fqk: Full Qualified Key to search for
    :param row_number: In which row to search for the fqk (1-based index)
    :return: column letter where the key can be found
    """

    # (for -> else) If the fqk is not found append a column creating it.
    for cell in sheet[row_number]:
        if cell.value == fqk:
            column_letter = get_column_letter(cell.column)
            break
    else:
        column_letter = get_column_letter(sheet.max_column + 1)
        sheet[f"{column_letter}{row_number}"].value = fqk
    return column_letter


def parse_spreadsheet(hca_spreadsheet: Workbook, entity_dictionary: Dict):
    """
    Parse the spreadsheet and fill the metadata with accessions.

    :param hca_spreadsheet: Workbook object of the spreadsheet
    :param entity_dictionary: Dictionary mapping by entity UUID to the proper archiveEntity
    :return: Accessioned spreadsheet
    """
    # Parse each sheet for the UUIDs
    for sheet in hca_spreadsheet.sheetnames:
        for row in hca_spreadsheet[sheet].rows:
            if row[0].value in entity_dictionary:
                # Get fqk, search for it, add accession based on the entity dictionary
                fqk = (accession_mapping[entity_dictionary[row[0].value]['type']]['fqk']
                       .replace("{}", sheet.lower().replace(" ", "_")))
                coordinate_column = search_fqk_in_sheet(hca_spreadsheet[sheet], fqk, 4)
                coordinate_row = row[0].coordinate[1:]
                cell_coordinate = f'{coordinate_column}{coordinate_row}'
                hca_spreadsheet[sheet][cell_coordinate].value = entity_dictionary[row[0].value]['accession']

    return hca_spreadsheet


def get_entity_dictionary(archive_entities_list: List[Dict]) -> Dict:
    """
    Given a list of ingest's archiveEntities, return a dictionary mapping UUIDs to the archiveEntities

    :param archive_entities_list: List of archiveEntities
    :return: Map of ingest UUIDs to archiveEntities metadata
    """

    entity_dictionary = {}

    for entity in archive_entities_list:
        # A single entity may refer to multiple files (e.g. 1 SequencingRun -> 3 data files)
        for accessionedMetadata in entity.get('accessionedMetadataUuids', []):
            entity_dictionary[accessionedMetadata] = entity

    return entity_dictionary


def get_all_entities(submission_uuid: str) -> List[Dict]:
    """
    Get all Ingest's archiveEntitites for a submission

    :param submission_uuid: UUID of the submission in ingest
    :return: List containing all archiveEntities for a submission
    """
    # Get the archiveEntities response and put them into a list
    archive_submission = rq.get(f'https://api.ingest.archive.data.humancellatlas.org/archiveSubmissions/search/'
                                f'findBySubmissionUuid?submissionUuid={submission_uuid}').json()
    entities_link = archive_submission['_links']['entities']['href']
    entities_page = rq.get(entities_link).json()
    all_entities = entities_page['_embedded']['archiveEntities']

    # Deal with pagination
    while entities_page.get('_links', {}).get('next'):
        entities_page = rq.get(entities_page['_links']['next']['href']).json()
        all_entities.extend(entities_page['_embedded']['archiveEntities'])

    return all_entities


def get_hca_spreadsheet(uuid: str) -> Tuple[Workbook, str]:
    """
    Get the HCA spreadsheet from an ingest submission UUID loaded as an openpyxl Workbook.

    :param uuid: Ingest submission UUID
    :return: Tuple(Workbook, filename)
    """

    url = f'https://ingest.archive.data.humancellatlas.org/submissions/{uuid}/spreadsheet'

    # Get spreadsheet from URL
    response = rq.get(url)
    response.raise_for_status()

    # Get openpyxl object from request
    file = BytesIO(response.content)
    spreadsheet = load_workbook(file)

    # Get filename from headers
    filename = response.headers['Content-Disposition'].split('filename=')[1]

    return spreadsheet, filename


def parse_args() -> argparse.Namespace:
    """
    Parse the arguments and return them.

    :return: Arguments from Command Line
    """
    parser = argparse.ArgumentParser()
    parser.add_argument('--hca_spreadsheet_output', '-o', help='Path where the HCA spreadsheet will be saved',
                        dest="output_path", default=os.getcwd(), type=str)
    parser.add_argument('--submission_uuid', '-u', help='uuid of the ingest submission', required=True,
                        type=str)
    return parser.parse_args()


def main(submission_uuid, output_path):
    hca_spreadsheet, filename = get_hca_spreadsheet(submission_uuid)
    all_entities = get_all_entities(submission_uuid)
    entity_dictionary = get_entity_dictionary(all_entities)
    # TODO Work way of adding sequencingExperiment entities. Can't be done by uuid.
    hca_spreadsheet = parse_spreadsheet(hca_spreadsheet, entity_dictionary)
    hca_spreadsheet.save(f"{output_path}/{filename.replace('.xlsx', '_accessions.xlsx')}")


if __name__ == '__main__':
    args = parse_args()
    main(args.submission_uuid, args.output_path)
