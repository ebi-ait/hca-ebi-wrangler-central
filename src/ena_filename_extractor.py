""" Ena filename extractor

This script takes a JSON file from the ENA as input and fills the
Sequence file sheet from the wrangling spreadsheet with the
corresponding values. It is assumed that the user has dowloaded the
JSON file from a project page at ENA (e.g. https://www.ebi.ac.uk/ena/data/view/PRJNA545828).
It is also assumed that the wrangling spreadsheet has filled the column
'INSDC EXPERIMENT ACCESSION (Required)' at the 'Sequence file' sheet.
In case the 'INPUT CELL SUSPENSION ID (Required)' column has not been
filled yet in the 'Sequence file sheet', the user can specify the
argument -cs. The script will then fill in that column. In this case,
it assumes that you have available a 'Cell suspension' sheet filled
with at leas the columns 'INSDC EXPERIMENT ACCESSION (Required)' and
'CELL SUSPENSION ID (Required)'.

The spreadsheets must be in .xlsx format.

This script requires the following packages:
    - pandas
    - json
    - openpyxl
    - argparse

The file can also be imported as a module and contains the following
functions:
    - ena_dict: create a dictionary with the fastq files per accession
    - fill_spreadsheet: fill the dataframe with the corresponding name of files
    - input_cell_suspension: fill the dataframe with the corresponding cell suspension ID
    - reorder_sheet: reorder the rows of a sheet based on column matching
"""

import json
import argparse
from os.path import isfile

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


def ena_dict(ena_report):
    """ Create a dictionary with the fastq files per accession.

    Creates a dictionary where the key is the run accession and
    the value is a list with the index of the run accession in the
    JSON structure, and the number of fastq files associated to it

    Args:
        ena_report (list): list of dictionaries from a JSON file

    Returns:
        ena_run_id (dict): dictionary with accessions as keys
                           and fastq files as values 
    """
    ena_run_id = {}
    for i, run in enumerate(ena_report):
        if not run.get('experiment_accession', '') in ena_run_id:
            ena_run_id[run.get('experiment_accession', '')] = [[], []]
        ena_run_id[run.get('experiment_accession')][0].append(i)
        ena_run_id[run.get('experiment_accession')][1].append(
            len(run.get('fastq_ftp', '').split(';')))

    if "" in ena_run_id:
        del ena_run_id[""]

    return ena_run_id


def fill_spreadsheet(ena_run_id, ena_report, new_df, original_df):
    """ Fill the dataframe with the corresponding name of files

    Add row by row the values in the same order as in the
    ENA run accession list. When it's a single file, copy the values
    in the original spreadsheet and complete it with the new file
    values. When it has two or more fastq files, repeat the process.

    Args:
        ena_run_id (dict): dictionary created by ena_dict
        ena_report (list): list of dictionaries from a JSON file
        new_df (pandas.dataframe): 'Sequence file' sheet as pandas dataframe
        original_df (pandas.dataframe): Wrangling spreadsheet as pandas
                                        dataframe, sequence file sheet

    Returns:
        new_df (pandas.dataframe): 'Sequence file' sheet updated
    """
    for key, values in ena_run_id.items():
        for i in range(len(values[0])):
            index_report = values[0][i]
            fastq_number = values[1][i]
            for j in range(fastq_number):
                new_df = new_df.append(
                        original_df.loc[original_df['process.insdc_experiment.insdc_experiment_accession'] == key])
                new_df.reset_index(drop=True, inplace=True)
                new_df.loc[new_df.index[-1], 'sequence_file.file_core.file_name'] = \
                    ena_report[index_report]['fastq_ftp'].split(';')[j].split('/')[-1]
                new_df.reset_index(drop=True, inplace=True)
            if not fastq_number:  # if the number of fastq files is 0
                print('No fastq files for run accession: ' + key)

    return new_df


def input_cell_suspension(original_df, suspension_worksheet):
    """ Fill the dataframe with the corresponding cell suspension ID

    Match the Cell suspension IDs with the Experiment Acession IDs.
    Write the Cell suspension IDs in the corresponding cells at the
    column Input Cell Suspension ID

    Args:
        original_df (pandas.dataframe): Cell suspension sheet as pandas
                                        dataframe
        suspension_worksheet (pandas.dataframe): Cell suspension sheet as pandas
                                    dataframe

    Returns:
        original_df (pandas.dataframe): Wrangling spreadsheet updated
    """

    for i in original_df.index:
        original_df.loc[i, 'cell_suspension.biomaterial_core.biomaterial_id'] = \
            suspension_worksheet.loc[suspension_worksheet['INSDC EXPERIMENT ACCESSION (Required)']
            == original_df.loc[i, 'process.insdc_experiment.insdc_experiment_accession'],
                                     'cell_suspension.biomaterial_core.biomaterial_id'].values[0]

    return original_df


def reorder_sheet(source_sheet, source_column, destination_sheet, destination_column):
    """Reorder the rows of a sheet based on column matching.

    Reorder the rows of destination_sheet according to the values of
    source_sheet in the column source_column, by matching those values
    with the rows of destination_sheet in the column destination_column.

    Args:
        source_sheet (pandas.dataframe): source sheet as pandas dataframe
        source_column (str): name of the column to be moved
        destination_sheet (pandas.dataframe): destination sheet as pandas dataframe
        destination_column (str): name of the column to be moved

    Returns:
        empty_pandas_df (pandas.dataframe): reordered destination_sheet
    """
    order = source_sheet[source_column].tolist()
    empty_pandas_df = pd.DataFrame(
        columns=destination_sheet.columns.values.tolist())
    for i in order:
        empty_pandas_df = empty_pandas_df.append(
            destination_sheet.loc[destination_sheet[destination_column] == i])
        empty_pandas_df.reset_index(drop=True, inplace=True)

    return empty_pandas_df


def main(args):
    # Load the JSON file
    with open(args.input) as json_file:
        ena_report = json.load(json_file)

    # Create the dictionary
    ena_run_id = ena_dict(ena_report)

    # Read the spreadsheet as a pandas dataframe
    pandas_worksheet = pd.read_excel(
        io=args.spreadsheet, sheet_name='Sequence file', header=3)
    # Drop the first 4 rows, which are not useful
    pandas_worksheet.drop(pandas_worksheet.index[:4], inplace=True)
    # Create a new empty dataframe with the same column names as the previous one
    df = pd.DataFrame(columns=pandas_worksheet.columns.values.tolist())

    # Fill the dataframe with the corresponding name of files
    df = fill_spreadsheet(ena_run_id, ena_report, df, pandas_worksheet)

    # If needed, fill the dataframe with the corresponding cell suspension ID
    # and reorder the spreadsheet to match the order of Cell suspension IDs
    # at the cell suspension sheet
    if args.cell_suspension == True:
        suspension_worksheet = pd.read_excel(
            io=args.spreadsheet, sheet_name='Cell suspension')
        suspension_worksheet.drop(suspension_worksheet.index[:4], inplace=True)
        df = input_cell_suspension(df, suspension_worksheet)

    # Save the dataframe to the destination spreadsheet
    # Load the spreadsheet
    workbook = load_workbook(args.spreadsheet)
    # Select the 'Sequence file' sheet as active
    ws = workbook['Sequence file']

    # Write the new dataframe to the spreadsheet taking into account
    # that the first 5 rows have to maintain the original information

    rows = dataframe_to_rows(df, index=False, header=False)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx+5, column=c_idx).value = value

    # Save the spreadsheet
    if args.output:
        workbook.save(args.output)
    else:
        workbook.save(args.spreadsheet)


#######################
# Main
#######################
if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='''\
        Convert ENA JSON to spreadsheet. This script takes as input a
        JSON file dowloaded from ENA with information about the experiment and the
        ftp links to dowload the files. It also needs as input/output a wrangling
        spreadsheet with at least one sheet named as "Sequence file". It will match
        the run accession IDs from the spreadsheet with those in the JSON, performing
        the following actions:
        1) Add extra entries for those runs in the spreadsheet with more than one file
        associated, copying the associated content
        2) Write the name of the files in the ftp link to the corresponding cells\n
        3) Save the updated spreadsheet
        ''')
    parser.add_argument(
        '-i', '--input', help='Input JSON file', type=isfile, required=True)
    parser.add_argument(
        '-s', '--spreadsheet', help='Wrangling spreadsheet, must be xlsx', type=isfile, required=True)
    parser.add_argument(
        '-o', '--output', help='Alternative output location, including xlsx filename', type=str, required=False)
    parser.add_argument(
        '-cs', '--cell-suspension', help='Include if you want to write and match the Cell suspension IDs with the Experiment Accession IDs',
        action='store_true', dest='cell_suspension')
    parser.set_defaults(cell_suspension=False)

    args = parser.parse_args()
    main(args)
